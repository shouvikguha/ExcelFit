import openpyxl as xl
import os

def create_workbook(filename):
    if not os.path.exists(filename):
        wb = xl.Workbook()
        # if len(wb.sheetnames) > 1:
        #     wb.remove(wb.active)
        wb.save(filename)

def add_data(filename, sheet_name, questions, data):
    wb = xl.load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        headers = ["Date & Time"] + questions
        sheet.append(headers)
    else:
        sheet = wb[sheet_name]

    sheet.append(data)
    wb.save(filename)